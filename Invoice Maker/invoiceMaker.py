import xlrd, datetime, re
import openpyxl

loc = "[Insert Input File Path here]"
w = openpyxl.load_workbook('[Insert Output File Path here]')
ws = w['Invoice']
ws['k9'] = int(ws['k9'].value) + 1
wb = xlrd.open_workbook(loc)
sheet = wb.sheet_by_index(0)

# For row 0 and column 0
sheet.cell_value(0, 0)

#covert XLS date float to actual date
temp = datetime.datetime(1899, 12, 30)
hello = str(temp + datetime.timedelta(days=sheet.cell_value(1, 1)))
value = re.findall(r"[\d]{1,4}-[\d]{1,2}-[\d]{1,2}", hello)
prev = value.pop()
count = -1
over = 0
z = 18
for i in range(1, sheet.nrows):
    var = sheet.cell_value(i, 1)
    var2 = sheet.cell_value(i, 5)
    hell = str(temp + datetime.timedelta(days=var))
    value = re.findall(r"[\d]{1,4}-[\d]{1,2}-[\d]{2}", hell)
    current = value.pop()
    count += 1
    if (var2 > 10):
        over += 1
    if prev != current:
        print()
        print()
        print(prev)
        ws['d{}'.format(z)] = prev
        ws['i{}'.format(z)] = count - over
        ws['i{}'.format(z + 1)] = over
        print("Total packages delivered: ", count)
        print("Overweight(11 or more lbs): ", over)
        prev = current
        count = 0
        over = 0
        z += 2
print()
print()
print(prev)
ws['k8'] = prev
ws['d{}'.format(z)] = prev
ws['i{}'.format(z)] = count + 1 - over
ws['i{}'.format(z + 1)] = over
print("Total packages delivered: ", count + 1)
print("Overweight(11 or more lbs): ", over)
print()
print()
print("Total packages: ", sheet.nrows - 1)

##########################################################################################################################

